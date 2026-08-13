"""Build a compact, formula-driven Excel analysis workbook from curated evidence."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo

from .config import EXCEL_DIR, PROCESSED_DIR, ensure_project_directories
from .utils import configure_logging


LOGGER = configure_logging()
WORKBOOK_PATH = EXCEL_DIR / "automotive_commercial_analysis.xlsx"

NAVY = "1F4E79"
BLUE = "5B9BD5"
PALE_BLUE = "DCE6F1"
ORANGE = "D97706"
GREEN = "548235"
RED = "C00000"
GREY = "667085"
PALE_GREY = "F2F4F7"
WHITE = "FFFFFF"
THIN_GREY = Side(style="thin", color="D0D5DD")


def _table(
    worksheet: object, name: str, reference: str, style: str = "TableStyleMedium2"
) -> None:
    table = Table(displayName=name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _title(worksheet: object, title: str, subtitle: str) -> None:
    worksheet.merge_cells("A1:H1")
    worksheet["A1"] = title
    worksheet["A1"].font = Font(size=20, bold=True, color=WHITE)
    worksheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    worksheet["A1"].alignment = Alignment(vertical="center")
    worksheet.row_dimensions[1].height = 34
    worksheet.merge_cells("A2:H2")
    worksheet["A2"] = subtitle
    worksheet["A2"].font = Font(size=10, italic=True, color=GREY)


def _header(row: Iterable[object]) -> None:
    for cell in row:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=THIN_GREY)


def _set_widths(worksheet: object, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _write_frame(
    worksheet: object,
    frame: pd.DataFrame,
    start_row: int,
    table_name: str,
) -> tuple[int, int]:
    headers = [str(column) for column in frame.columns]
    for column_index, value in enumerate(headers, start=1):
        worksheet.cell(start_row, column_index, value)
    _header(worksheet[start_row])
    for row_index, record in enumerate(
        frame.itertuples(index=False, name=None), start_row + 1
    ):
        for column_index, value in enumerate(record, start=1):
            if pd.isna(value):
                value = None
            worksheet.cell(row_index, column_index, value)
    end_row = start_row + len(frame)
    end_column = len(headers)
    _table(
        worksheet,
        table_name,
        f"A{start_row}:{worksheet.cell(start_row, end_column).column_letter}{end_row}",
    )
    return end_row, end_column


def _source_sheet(workbook: Workbook, summary: dict[str, object]) -> object:
    sheet = workbook.create_sheet("KPI_Source")
    sheet.append(["metric_name", "value"])
    for name, value in summary["kpis"].items():
        sheet.append([name, value])
    _table(sheet, "tblKpiSource", f"A1:B{sheet.max_row}", "TableStyleMedium4")
    sheet.sheet_state = "hidden"
    return sheet


def _executive_sheet(workbook: Workbook, summary: dict[str, object]) -> object:
    sheet = workbook.active
    sheet.title = "Executive Summary"
    _title(
        sheet,
        "Automotive Commercial Analytics",
        "Decision workbook | Synthetic Swiss marketplace data | CHF | 2022-2025",
    )
    kpi_cards = [
        ("Total inquiries", "=KPI_Source!B2", "#,##0"),
        ("Conversion rate", "=KPI_Source!B5", "0.0%"),
        ("Accepted value (CHF m)", "=KPI_Source!B7", '"CHF" #,##0.0,,'),
        ("Quote anomaly rate", "=KPI_Source!B15", "0.00%"),
    ]
    for index, (label, formula, number_format) in enumerate(kpi_cards):
        start_column = 1 + index * 2
        sheet.merge_cells(
            start_row=4,
            start_column=start_column,
            end_row=4,
            end_column=start_column + 1,
        )
        sheet.cell(4, start_column, label)
        sheet.cell(4, start_column).font = Font(bold=True, color=GREY)
        sheet.merge_cells(
            start_row=5,
            start_column=start_column,
            end_row=6,
            end_column=start_column + 1,
        )
        value_cell = sheet.cell(5, start_column, formula)
        value_cell.font = Font(size=18, bold=True, color=NAVY)
        value_cell.fill = PatternFill("solid", fgColor=PALE_BLUE)
        value_cell.alignment = Alignment(vertical="center")
        value_cell.number_format = number_format

    sheet["A8"] = "Decision"
    sheet["B8"] = "Evidence"
    sheet["C8"] = "Management action"
    sheet["D8"] = "Owner"
    for merged_range in ["D8:H8"]:
        sheet.merge_cells(merged_range)
    _header(sheet[8])
    evidence = summary["evidence"]
    decisions = [
        (
            "Dealer response SLA",
            f"<8h conversion {evidence['conversion_under_8_hours']:.2%} vs {evidence['conversion_8_hours_or_more']:.2%}",
            "Pilot an 8-hour first-quote SLA; track conversion and dealer coverage.",
            "Marketplace operations",
        ),
        (
            "Channel allocation",
            f"Dealer Referral leads Social Media by {evidence['dealer_referral_lift_pp_vs_social']:.2f} pp",
            "Protect referral capacity; review social targeting and lead qualification.",
            "Commercial / marketing",
        ),
        (
            "Older-vehicle pricing",
            f"10+ year relative quote spread is {evidence['older_to_newer_relative_spread_ratio']:.2f}x the 0-2 year spread",
            "Set wider customer expectations and request stronger condition evidence.",
            "Pricing operations",
        ),
        (
            "Pricing exceptions",
            f"{summary['kpis']['quote_anomalies']:,} soft anomalies remain reviewable",
            "Triage flagged quotes; keep hard-invalid values out of reporting.",
            "Data + pricing",
        ),
    ]
    for row_number, decision in enumerate(decisions, start=9):
        for column_number, value in enumerate(decision, start=1):
            sheet.cell(row_number, column_number, value)
        sheet.merge_cells(
            start_row=row_number, start_column=4, end_row=row_number, end_column=8
        )
        for cell in sheet[row_number]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(bottom=THIN_GREY)
        sheet.row_dimensions[row_number].height = 42

    sheet["A15"] = (
        "Use the detailed tabs to challenge each headline. Workbook formulas recalculate in Excel; source values come from analysis_summary.json."
    )
    sheet.merge_cells("A15:H15")
    sheet["A15"].font = Font(italic=True, color=GREY)
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    _set_widths(
        sheet,
        {"A": 24, "B": 24, "C": 66, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14},
    )
    return sheet


def _monthly_sheet(workbook: Workbook, summary: dict[str, object]) -> object:
    sheet = workbook.create_sheet("Monthly Performance")
    _title(
        sheet, "Monthly Performance", "Inquiry-grain trend with formula-derived rates"
    )
    frame = pd.DataFrame(summary["monthly"])[
        ["month", "inquiries", "conversions", "accepted_value"]
    ]
    frame["month"] = pd.to_datetime(frame["month"]).dt.date
    frame.columns = ["Month", "Inquiries", "Conversions", "Accepted Value (CHF m)"]
    frame.insert(3, "Conversion Rate", None)
    frame.insert(5, "MoM Inquiry Growth", None)
    end_row, _ = _write_frame(sheet, frame, 4, "tblMonthlyPerformance")
    for row in range(5, end_row + 1):
        sheet.cell(row, 4, f"=IFERROR(C{row}/B{row},0)")
        sheet.cell(row, 6, "" if row == 5 else f'=IFERROR(B{row}/B{row-1}-1,"")')
        sheet.cell(row, 1).number_format = "mmm-yy"
        sheet.cell(row, 4).number_format = "0.0%"
        sheet.cell(row, 5).number_format = '"CHF" #,##0.0,,'
        sheet.cell(row, 6).number_format = "+0.0%;-0.0%;0.0%"
    sheet.conditional_formatting.add(
        f"D5:D{end_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    line = LineChart()
    line.title = "Monthly inquiries and conversion"
    line.y_axis.title = "Inquiries"
    line.x_axis.title = "Month"
    line.add_data(
        Reference(sheet, min_col=2, min_row=4, max_row=end_row), titles_from_data=True
    )
    line.set_categories(Reference(sheet, min_col=1, min_row=5, max_row=end_row))
    rate = LineChart()
    rate.add_data(
        Reference(sheet, min_col=4, min_row=4, max_row=end_row), titles_from_data=True
    )
    rate.y_axis.axId = 200
    rate.y_axis.title = "Conversion rate"
    rate.y_axis.numFmt = "0%"
    rate.y_axis.crosses = "max"
    line += rate
    line.height = 8
    line.width = 16
    sheet.add_chart(line, "H4")
    sheet.freeze_panes = "A5"
    _set_widths(sheet, {"A": 13, "B": 13, "C": 13, "D": 17, "E": 18, "F": 19})
    return sheet


def _lead_source_sheet(workbook: Workbook, summary: dict[str, object]) -> object:
    sheet = workbook.create_sheet("Lead Source")
    _title(sheet, "Lead Source", "Volume, conversion, value, and response trade-offs")
    frame = pd.DataFrame(summary["lead_sources"])[
        [
            "lead_source",
            "inquiries",
            "conversions",
            "average_sale_value",
            "fastest_response_hours",
        ]
    ]
    frame.columns = [
        "Lead Source",
        "Inquiries",
        "Conversions",
        "Average Sale Value",
        "Average Fastest Response (h)",
    ]
    frame.insert(3, "Conversion Rate", None)
    frame["Portfolio Lift (pp)"] = None
    frame["Conversion Rank"] = None
    end_row, _ = _write_frame(sheet, frame, 4, "tblLeadSource")
    total_inquiries = f"SUM(B5:B{end_row})"
    total_conversions = f"SUM(C5:C{end_row})"
    for row in range(5, end_row + 1):
        sheet.cell(row, 4, f"=IFERROR(C{row}/B{row},0)")
        sheet.cell(row, 7, f"=(D{row}-({total_conversions}/{total_inquiries}))*100")
        sheet.cell(row, 8, f"=RANK.EQ(D{row},$D$5:$D${end_row},0)")
        sheet.cell(row, 4).number_format = "0.0%"
        sheet.cell(row, 5).number_format = '"CHF" #,##0'
        sheet.cell(row, 6).number_format = '0.0 "h"'
        sheet.cell(row, 7).number_format = '+0.0 "pp";-0.0 "pp";0.0 "pp"'
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Conversion rate by lead source"
    chart.add_data(
        Reference(sheet, min_col=4, min_row=4, max_row=end_row), titles_from_data=True
    )
    chart.set_categories(Reference(sheet, min_col=1, min_row=5, max_row=end_row))
    chart.height = 8
    chart.width = 15
    sheet.add_chart(chart, "J4")
    sheet.freeze_panes = "A5"
    _set_widths(
        sheet, {"A": 22, "B": 12, "C": 13, "D": 16, "E": 19, "F": 27, "G": 18, "H": 17}
    )
    return sheet


def _dealer_sheet(workbook: Workbook, summary: dict[str, object]) -> object:
    sheet = workbook.create_sheet("Dealer Performance")
    _title(
        sheet,
        "Dealer Performance",
        "Dealer quote grain; rank is volume-guarded at 250 quotes",
    )
    frame = pd.DataFrame(summary["dealer_detail"])[
        [
            "dealer_id",
            "dealer_name",
            "dealer_region",
            "quotes",
            "wins",
            "average_response_hours",
            "average_quote_to_market_ratio",
            "accepted_value",
        ]
    ]
    frame.columns = [
        "Dealer ID",
        "Dealer",
        "Region",
        "Quotes",
        "Wins",
        "Average Response (h)",
        "Average Quote / Market",
        "Accepted Value",
    ]
    frame.insert(5, "Win Rate", None)
    frame["Eligible"] = None
    frame["Eligible Rank"] = None
    end_row, _ = _write_frame(sheet, frame, 4, "tblDealerPerformance")
    for row in range(5, end_row + 1):
        sheet.cell(row, 6, f"=IFERROR(E{row}/D{row},0)")
        sheet.cell(row, 10, f'=IF(D{row}>=250,"Yes","No")')
        sheet.cell(
            row, 11, f'=IF(J{row}="Yes",RANK.EQ(F{row},$F$5:$F${end_row},0),NA())'
        )
        sheet.cell(row, 6).number_format = "0.0%"
        sheet.cell(row, 7).number_format = '0.0 "h"'
        sheet.cell(row, 8).number_format = "0.0%"
        sheet.cell(row, 9).number_format = '"CHF" #,##0'
    sheet.conditional_formatting.add(
        f"F5:F{end_row}", DataBarRule(start_type="min", end_type="max", color=GREEN)
    )
    scatter = ScatterChart()
    scatter.title = "Response time vs win rate"
    scatter.x_axis.title = "Average response hours"
    scatter.y_axis.title = "Win rate"
    scatter.series.append(
        Series(
            Reference(sheet, min_col=6, min_row=5, max_row=end_row),
            Reference(sheet, min_col=7, min_row=5, max_row=end_row),
            title="Dealers",
        )
    )
    scatter.height = 8
    scatter.width = 15
    sheet.add_chart(scatter, "M4")
    sheet.freeze_panes = "A5"
    _set_widths(
        sheet,
        {
            "A": 11,
            "B": 28,
            "C": 22,
            "D": 11,
            "E": 10,
            "F": 12,
            "G": 21,
            "H": 22,
            "I": 17,
            "J": 11,
            "K": 14,
        },
    )
    return sheet


def _quality_sheet(workbook: Workbook) -> object:
    sheet = workbook.create_sheet("Data Quality")
    _title(
        sheet,
        "Data Quality",
        "Comparable checks at raw and processed stages; score is check-weighted",
    )
    metrics = pd.read_csv(PROCESSED_DIR / "data_quality_metrics.csv")[
        ["stage", "metric_name", "passed_records", "total_records", "description"]
    ]
    metrics.columns = [
        "Stage",
        "Metric",
        "Passed Checks",
        "Applicable Checks",
        "Description",
    ]
    metrics.insert(4, "Rate", None)
    end_row, _ = _write_frame(sheet, metrics, 4, "tblDataQuality")
    for row in range(5, end_row + 1):
        sheet.cell(row, 5, f"=IFERROR(C{row}/D{row},1)")
        sheet.cell(row, 5).number_format = "0.000%"
    sheet.conditional_formatting.add(
        f"E5:E{end_row}",
        ColorScaleRule(
            start_type="min",
            start_color="F8696B",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFEB84",
            end_type="max",
            end_color="63BE7B",
        ),
    )
    issue_start = end_row + 3
    sheet.cell(issue_start - 1, 1, "Issue log (treatment evidence)").font = Font(
        bold=True, color=NAVY, size=13
    )
    issues = pd.read_csv(PROCESSED_DIR / "data_quality_issue_log.csv")
    issues.columns = [column.replace("_", " ").title() for column in issues.columns]
    _write_frame(sheet, issues, issue_start, "tblQualityIssues")
    sheet.freeze_panes = "A5"
    _set_widths(sheet, {"A": 15, "B": 34, "C": 16, "D": 18, "E": 14, "F": 65})
    return sheet


def _dictionary_sheet(workbook: Workbook) -> object:
    sheet = workbook.create_sheet("Field Profile")
    _title(
        sheet,
        "Curated Field Profile",
        "Machine-derived schema, nullability, cardinality, and examples",
    )
    rows: list[dict[str, object]] = []
    for path in sorted(PROCESSED_DIR.glob("*.csv")):
        if path.stem in {"analysis_kpis"}:
            continue
        frame = pd.read_csv(path, low_memory=False)
        for column in frame.columns:
            non_null = frame[column].dropna()
            rows.append(
                {
                    "Table": path.stem,
                    "Column": column,
                    "Pandas Type": str(frame[column].dtype),
                    "Rows": len(frame),
                    "Null Rate": float(frame[column].isna().mean()),
                    "Distinct Values": int(non_null.nunique()),
                    "Example": None if non_null.empty else str(non_null.iloc[0])[:80],
                }
            )
    profile = pd.DataFrame(rows)
    end_row, _ = _write_frame(sheet, profile, 4, "tblFieldProfile")
    for row in range(5, end_row + 1):
        sheet.cell(row, 5).number_format = "0.00%"
    sheet.freeze_panes = "A5"
    _set_widths(sheet, {"A": 28, "B": 31, "C": 14, "D": 12, "E": 13, "F": 16, "G": 40})
    return sheet


def validate_excel_workbook(path: Path = WORKBOOK_PATH) -> dict[str, int]:
    """Reopen the workbook and assert its recruiter-facing contract."""

    workbook = load_workbook(path, data_only=False, read_only=False)
    required = {
        "Executive Summary",
        "Monthly Performance",
        "Lead Source",
        "Dealer Performance",
        "Data Quality",
        "Field Profile",
        "KPI_Source",
    }
    if not required.issubset(workbook.sheetnames):
        raise ValueError(
            f"Workbook is missing sheets: {sorted(required - set(workbook.sheetnames))}"
        )
    formula_cells = sum(
        1
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )
    error_cells = [
        f"{sheet.title}!{cell.coordinate}"
        for sheet in workbook.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.data_type == "e"
    ]
    chart_count = sum(len(sheet._charts) for sheet in workbook.worksheets)
    if formula_cells < 50 or chart_count < 3 or error_cells:
        raise ValueError(
            "Workbook contract requires at least 50 formulas, 3 charts, and no error cells"
        )
    return {
        "sheets": len(workbook.sheetnames),
        "formulas": formula_cells,
        "charts": chart_count,
    }


def build_excel_workbook(path: Path = WORKBOOK_PATH) -> Path:
    """Build, save, and reopen a management-ready workbook."""

    ensure_project_directories()
    summary_path = PROCESSED_DIR / "analysis_summary.json"
    if not summary_path.exists():
        raise FileNotFoundError("Run `python -m src.analysis` before building Excel")
    summary = json.loads(summary_path.read_text(encoding="utf-8"))

    workbook = Workbook()
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    _executive_sheet(workbook, summary)
    _monthly_sheet(workbook, summary)
    _lead_source_sheet(workbook, summary)
    _dealer_sheet(workbook, summary)
    _quality_sheet(workbook)
    _dictionary_sheet(workbook)
    _source_sheet(workbook, summary)

    workbook.properties.creator = "Automotive Commercial Analytics"
    workbook.properties.title = "Swiss Automotive Commercial Decision Workbook"
    workbook.properties.subject = "Synthetic marketplace KPI and data-quality analysis"
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    evidence = validate_excel_workbook(path)
    LOGGER.info(
        "Built %s (%s sheets, %s formulas, %s charts)",
        path,
        evidence["sheets"],
        evidence["formulas"],
        evidence["charts"],
    )
    return path


def main() -> None:
    """CLI entry point."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=WORKBOOK_PATH)
    args = parser.parse_args()
    build_excel_workbook(args.output)


if __name__ == "__main__":
    main()
